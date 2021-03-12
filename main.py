import requests
import re
import ipaddress
import pandas as pd
import openpyxl
from tkinter import * 
from tkinter import filedialog
import tkinter.messagebox
import os
from openpyxl.styles import Border, Side
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

apifile = open("API_KEY.ini", "r+")
api_key = apifile.read()
apifile.close()

filepath = None

GUI = Tk()

def checkip(IP):
    if ipaddress.ip_address(IP).is_private is False:
        abipdbheaders = {
            'Key': api_key,
            'Accept': 'application/json',
        }

        abipdbparams = {
            'maxAgeInDays': 1,
            'ipAddress': IP,
        }

        req = requests.get("https://api.abuseipdb.com/api/v2/check", headers = abipdbheaders, params = abipdbparams)
        resp = req.json()

        if 'errors' not in resp:
            return resp["data"]
        else:
            exit()
    else:
        return (f"{IP} is private")

def filterip(ipin) :
    ipregex = re.compile(r'(?:\d{1,3}\.)+(?:\d{1,3})') 
    ipa = re.search(ipregex, ipin)
    return ipa.group(0)

def checkipfromfile(infile):
    iplist = []
    output = []
    f1 = open(infile, 'r')
    tmp = f1.readlines()
    
    for i in tmp:   
        if i == '' or i == " " or i == "198.19.250.1" or i == "\n":
            pass
        else: 
            iplist.append(filterip(i))
    for i in iplist:
        output.append(checkip(i))
    f1.close()
    
    return output

def checkipfrominput(in1):
    iplist2 = []
    output2 = []

    tmp2 = in1.split('\n')

    for i in tmp2:   
        if i == '' or i == " " or i == "198.19.250.1" or i == "\n":
            pass
        else: 
            iplist2.append(filterip(i))
    for i in iplist2:
        output2.append(checkip(i))

    return output2

def get_report(input):
    concdict = {
    k: [d.get(k) for d in input if k in d]
    for k in set().union(*input)
}

    IpaddList =  concdict.get("ipAddress")
    AbusescoreList = concdict.get("abuseConfidenceScore")
    PublicList = concdict.get("isPublic")
    IpverList = concdict.get("ipVersion")
    IswlList = concdict.get("isWhitelisted")
    CountrycList = concdict.get("countryCode")
    UsageList = concdict.get("usageType")
    IspList = concdict.get("isp")
    DomainList = concdict.get("domain")
    TotalreportsList = concdict.get("totalReports")
    LastreportList = concdict.get("lastReportedAt")

    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws['A1'] = 'ipAddress'
    ws['B1'] = 'abuseConfidenceScore'
    ws['C1'] = 'isPublic'
    ws['D1'] = 'ipVersion'
    ws['E1'] = 'isWhitelisted'
    ws['F1'] = 'countryCode'
    ws['G1'] = 'usageType'
    ws['H1'] = 'isp'
    ws['I1'] = 'domain'
    ws['J1'] = 'totalReports'
    ws['K1'] = 'lastReportedAt'

    border_style = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    clrrule = ColorScaleRule(start_type= 'num', start_value='0',start_color='00B050', mid_type= 'num', mid_value='25', mid_color='FCA904', end_type='num', end_value='100', end_color='CC0000')
    ws.conditional_formatting.add('B2:B500', clrrule)
    ws.conditional_formatting.add('A1:K500', FormulaRule(formula=['NOT(ISBLANK(A1))'], stopIfTrue=False, border=border_style))

    dataframeIpaddList = pd.DataFrame({'ipAddress': IpaddList})
    for index, row in dataframeIpaddList.iterrows():
        cell = 'A%d'  % (index + 2)
        ws[cell] = row[0]
    dataframeAbusescoreList = pd.DataFrame({'abuseConfidenceScore': AbusescoreList})
    for index, row in dataframeAbusescoreList.iterrows():
        cell = 'B%d'  % (index + 2)
        ws[cell] = row[0]
    dataframePublicList = pd.DataFrame({'isPublic': PublicList})
    for index, row in dataframePublicList.iterrows():
        cell = 'C%d'  % (index + 2)
        ws[cell] = row[0]
    dataframeIpverList = pd.DataFrame({'ipVersion': IpverList})
    for index, row in dataframeIpverList.iterrows():
        cell = 'D%d'  % (index + 2)
        ws[cell] = row[0]
    dataframeIswlList = pd.DataFrame({'isWhitelisted': IswlList})
    for index, row in dataframeIswlList.iterrows():
        cell = 'E%d'  % (index + 2)
        ws[cell] = row[0]
    dataframeCountrycList = pd.DataFrame({'countryCode': CountrycList})
    for index, row in dataframeCountrycList.iterrows():
        cell = 'F%d'  % (index + 2)
        ws[cell] = row[0]
    dataframeUsageList = pd.DataFrame({'usageType': UsageList})
    for index, row in dataframeUsageList.iterrows():
        cell = 'G%d'  % (index + 2)
        ws[cell] = row[0]
    dataframeIspList = pd.DataFrame({'isp': IspList})
    for index, row in dataframeIspList.iterrows():
        cell = 'H%d'  % (index + 2)
        ws[cell] = row[0]
    dataframeDomainList = pd.DataFrame({'domain': DomainList})
    for index, row in dataframeDomainList.iterrows():
        cell = 'I%d'  % (index + 2)
        ws[cell] = row[0]
    dataframeTotalreportsList = pd.DataFrame({'totalReports': TotalreportsList})
    for index, row in dataframeTotalreportsList.iterrows():
        cell = 'J%d'  % (index + 2)
        ws[cell] = row[0]
    dataframeLastreportList = pd.DataFrame({'lastReportedAt': LastreportList})
    for index, row in dataframeLastreportList.iterrows():
        cell = 'K%d'  % (index + 2)
        ws[cell] = row[0]

    dims = {}
    for i in ws.rows:
        for cell in i:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
    for x, y in dims.items():
        ws.column_dimensions[x].width = y

    wb.save("results.xlsx")

def selectfile():
    global filepath
    filepath = filedialog.askopenfilename()
    filename = os.path.basename(filepath)
    pathlabel.config(text= "File Selected:\n" + os.path.basename(filename))

def removefile():
    global filepath
    filepath = None 
    pathlabel.config(text="No File Selected.\n")

def removetext():
    e1.delete("0.0", END)

def run(manual):
    try:
        if manual != '\n' and filepath is None:
            if api_key != "" or api_key != " ":
                get_report(checkipfrominput(manual))
            else:
                tkinter.messagebox.showinfo(title="AbuseIPDB Scanner", message="Error. No API Key Found.")
                GUI.destroy
                exit()
            
            msgboxresult = tkinter.messagebox.askyesno(title="AbuseIPDB Scanner", message="Success! Would you like to see Results?")
            if msgboxresult == True:
                os.system("start EXCEL.EXE results.xlsx")
                GUI.destroy()
            else:
                GUI.destroy()
        elif filepath is not None and manual == '\n':            
            if api_key != "" or api_key != " ":
                get_report(checkipfromfile(filepath))
            else:
                tkinter.messagebox.showinfo(title="AbuseIPDB Scanner", message="Error. No API Key Found.")
                GUI.destroy
                exit()
            
            msgboxresult = tkinter.messagebox.askyesno(title="AbuseIPDB Scanner", message="Success! Would you like to see Results?")
            if msgboxresult == True:
                os.system("start EXCEL.EXE results.xlsx")
                GUI.destroy()
            else:
                GUI.destroy()
        else:
            msgboxresult = tkinter.messagebox.showerror(title="AbuseIPDB Scanner", message="An Error occured.\nPlease ensure manual box is empty if is used")
        
    except Exception as e:
            msgboxresult = tkinter.messagebox.showerror(title="AbuseIPDB Scanner", message=f"An Error occured.\nPlease Try Again.\n({e})")

GUI.title("AbuseIPDB Scanner")
GUI.configure(background="#4b4c4c")
GUI.minsize(350,300)
GUI.resizable(0,0)

lbl1 = Label(GUI, text= "AbuseIPDB Scanner\n", bg= "#4b4c4c", fg= "red", font= "none 14 bold ")
lbl1.config(anchor=CENTER)
lbl1.pack(padx = 0, pady = 10)

lbl2 = Label(GUI, text= "Please Select IP Address File.\n", bg= "#4b4c4c", fg= "white", font= "none 10 ")
lbl2.config(anchor=CENTER)
lbl2.pack(padx = 0, pady = 0)

pathlabel = Label(GUI, text="No File Selected.\n", bg= "#4b4c4c", fg= "white", font= "Arial 8 bold")
pathlabel.config(anchor=CENTER)
pathlabel.pack(padx = 0, pady = 10)

btn1 = Button(GUI, text="Browse File",  command=lambda: [selectfile()], bg= "#a9a9a9", fg= "white", font= "Arial 8 bold")
btn1.config(anchor=CENTER)
btn1.pack(padx = 0, pady = 10)

btn3 = Button(GUI, text="Remove", command=lambda: [removefile()], bg= "#a9a9a9", fg= "white", font= "Arial 8 bold")
btn3.config(anchor=CENTER)
btn3.pack(padx = 0, pady = 0)

lbl3 = Label(GUI, text= "Or Manually Input IP Addresses.\n(1 IP Address Per line)", bg= "#4b4c4c", fg= "white", font= "none 10 ")
lbl3.config(anchor=CENTER)
lbl3.pack(padx = 0, pady = 0)

e1 = Text(GUI, height= 15,width = 30)
e1.insert(END, '')
e1.pack(padx = 0, pady = 10)

btn4 = Button(GUI, text="Remove", command=lambda: [removetext()], bg= "#a9a9a9", fg= "white", font= "Arial 8 bold")
btn4.config(anchor=CENTER)
btn4.pack(padx = 0, pady = 0)

btn2 = Button(GUI, text = "RUN",  command=lambda: [run(e1.get("0.0", END))], bg= "white", fg= "black", font= "Arial 10 bold" )
btn2.config(anchor=CENTER)
btn2.pack(padx = 0, pady = 20)

GUI.mainloop()
